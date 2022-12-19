# First we need to install openpyxl module to open an Excel sheet

import openpyxl
path = "/Users/akulac/PycharmProjects/Selenium project/Excel/SELENIUMLEARN.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active   # or we can also use  #workbook.get sheet by name ("Sheet]")

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)